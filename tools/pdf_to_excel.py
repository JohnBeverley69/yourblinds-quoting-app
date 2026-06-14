#!/usr/bin/env python3
"""
pdf_to_excel.py — convert a supplier price-list PDF into an Excel workbook that
the YourBlinds supplier importer can read.

Many suppliers ship price lists as (text-based) PDFs. The app's importer reads
spreadsheets, not PDFs, so this is the bridge: it extracts each price-table page
with pdfplumber and writes it to its own worksheet (named after the product),
laid out so the existing band-block parser can pick it up.

It's an OFFLINE helper (run on a PC with Python) so the live PHP server stays
simple — you only re-import a supplier once or twice a year.

Usage:
    py pdf_to_excel.py  "input price list.pdf"  ["output.xlsx"]

If the output path is omitted it writes "<input>.converted.xlsx" alongside the
PDF. It is read-only on the PDF and never touches the app or its database.
"""

import re
import sys
import os

import pdfplumber
from openpyxl import Workbook

# Lines that appear on every page (headers/footers/boilerplate) and are NOT the
# product name. Used to find the real product title on a price page.
BOILER = re.compile(
    r"^(trade price list|issued|all prices|service and|deliveries|notes\b|www\.|"
    r"useful email|general enq|please|using the|order via|website ordering|"
    r"spare parts|out of stock|statement|imagery|unsubmitted|express|same day|"
    r"selected products|product with|international freephone|tel\b|email|"
    r"\d+[.)]\s|page\s+\d+)",
    re.I,
)

BAND_RE = re.compile(r"\bBand\s+([A-Z]{1,3})\b")


def clean_sheet_name(name, used):
    """Excel sheet names: <=31 chars, no \\ / ? * [ ] : and must be unique."""
    name = re.sub(r"[\\/?*\[\]:]", " ", name).strip() or "Sheet"
    name = name[:28]
    base, n = name, 2
    while name.lower() in used:
        suffix = " %d" % n
        name = base[: 28 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def product_title(text):
    """Best-effort product name = first meaningful (non-boilerplate) line."""
    for line in (text or "").splitlines():
        s = line.strip()
        if len(s) < 3:
            continue
        if BOILER.match(s):
            continue
        if not re.search(r"[A-Za-z]", s):
            continue
        # Trim a trailing "All prices are in ..." tail some headers carry.
        s = re.split(r"\s+All prices are in", s)[0].strip()
        return s[:60]
    return None


def largest_table(tables):
    """Pick the table with the most cells (the price grid, usually)."""
    best, best_cells = None, 0
    for t in tables or []:
        cells = sum(1 for row in t for c in row if (c or "").strip())
        if cells > best_cells:
            best, best_cells = t, cells
    return best, best_cells


def convert(pdf_path, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    pages_with_grid = 0
    sheets_written = 0

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for pi, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            table, cells = largest_table(page.extract_tables())
            # A price grid: a table with a decent number of cells and several
            # price-looking numbers in the page text.
            prices = len(re.findall(r"\d+\.\d{2}", text)) + len(re.findall(r"\b\d{2,4}\b", text))
            if table is None or cells < 12 or prices < 12:
                continue
            pages_with_grid += 1

            name = product_title(text) or ("Page %d" % (pi + 1))
            bands = BAND_RE.findall(text)
            single_band = bands[0] if len(set(bands)) == 1 else None

            sheet = wb.create_sheet(clean_sheet_name(name, used))
            r = 1
            sheet.cell(r, 1, name); r += 1
            # A single-band page gets a "Band X" header so the band-block parser
            # latches on. (Multi-band / matrix pages are written as-is and may
            # need a per-supplier profile to import.)
            if single_band:
                sheet.cell(r, 1, "Band %s" % single_band); r += 1
            r += 1  # spacer row

            for row in table:
                for ci, val in enumerate(row, start=1):
                    v = (val or "").replace("\n", " ").strip()
                    if v:
                        sheet.cell(r, ci, v)
                r += 1
            sheets_written += 1

    wb.save(out_path)
    return total, pages_with_grid, sheets_written


def main():
    if len(sys.argv) < 2:
        print("Usage: py pdf_to_excel.py \"input.pdf\" [\"output.xlsx\"]")
        sys.exit(1)
    pdf_path = sys.argv[1]
    if not os.path.isfile(pdf_path):
        print("File not found:", pdf_path)
        sys.exit(1)
    out_path = sys.argv[2] if len(sys.argv) > 2 else (os.path.splitext(pdf_path)[0] + ".converted.xlsx")

    print("Reading:", pdf_path)
    total, grids, sheets = convert(pdf_path, out_path)
    print("Pages: %d  |  price-grid pages: %d  |  product sheets written: %d" % (total, grids, sheets))
    print("Wrote:", out_path)
    print("Now upload that .xlsx in Master Admin -> Supplier import to preview it.")


if __name__ == "__main__":
    main()
